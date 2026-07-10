"""
Billable Hours Tracker
- Sends weekly SMS to consultants asking for billable hours
- Receives replies and stores them in SQLite
- Provides a dashboard to view/export data
"""

import csv
from calendar import monthrange
import io
import os
import re
import secrets
import sqlite3
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"), override=False)

from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, Response, jsonify, render_template, request, send_file
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from twilio.base.exceptions import TwilioRestException
from twilio.request_validator import RequestValidator
from twilio.rest import Client
from twilio.twiml.messaging_response import MessagingResponse

try:
    import sentry_sdk
    from sentry_sdk.integrations.flask import FlaskIntegration
except ImportError:
    sentry_sdk = None
    FlaskIntegration = None

app = Flask(__name__)

MESSAGE_TYPE_LABELS = {
    "weekly_prompt": "Weekly prompt",
    "reminder": "Reminder",
    "manual_follow_up": "Manual follow-up",
    "response": "Consultant reply",
}

# ─── Config (set these as environment variables) ───────────────────────────────
TWILIO_ACCOUNT_SID = os.environ.get("TWILIO_ACCOUNT_SID")
TWILIO_AUTH_TOKEN  = os.environ.get("TWILIO_AUTH_TOKEN")
TWILIO_PHONE       = os.environ.get("TWILIO_PHONE")   # e.g. "+16175551234"
TWILIO_MESSAGING_SERVICE_SID = os.environ.get("TWILIO_MESSAGING_SERVICE_SID")
DB_PATH            = os.environ.get("DB_PATH", "hours.db")
PORT               = int(os.environ.get("PORT", "8000"))
CENTRAL_TIME       = ZoneInfo("America/Chicago")
ADMIN_USERNAME     = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD     = os.environ.get("ADMIN_PASSWORD")
VALIDATE_TWILIO_WEBHOOKS = os.environ.get("VALIDATE_TWILIO_WEBHOOKS", "false").lower() == "true"
SENTRY_DSN         = os.environ.get("SENTRY_DSN")
REPORT_OUTPUT_DIR  = os.environ.get("REPORT_OUTPUT_DIR", "reports")
PUBLIC_BASE_URL    = os.environ.get("PUBLIC_BASE_URL")
WEEKLY_PROMPT_UTC_HOUR = int(os.environ.get("WEEKLY_PROMPT_UTC_HOUR", "15"))
MONTHLY_REPORT_UTC_HOUR = int(os.environ.get("MONTHLY_REPORT_UTC_HOUR", "23"))

if SENTRY_DSN and sentry_sdk and FlaskIntegration:
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        integrations=[FlaskIntegration()],
        traces_sample_rate=float(os.environ.get("SENTRY_TRACES_SAMPLE_RATE", "0")),
    )

# ─── Helpers ───────────────────────────────────────────────────────────────────

def current_week_monday():
    """Return the ISO date string for the active reporting cycle Monday."""
    configured_cycle = get_reporting_cycle_date()
    if configured_cycle:
        return configured_cycle
    today = datetime.now(timezone.utc)
    monday = today - timedelta(days=today.weekday())
    return monday.strftime("%Y-%m-%d")

def validate_reporting_cycle_date(value: str):
    """Validate a YYYY-MM-DD date and normalize it to that week's Monday."""
    if not value:
        raise ValueError("reporting_cycle_date is required")
    parsed = datetime.strptime(value, "%Y-%m-%d")
    monday = parsed - timedelta(days=parsed.weekday())
    return monday.strftime("%Y-%m-%d")

def reporting_period_parts(week_of: str):
    parsed = datetime.strptime(week_of, "%Y-%m-%d")
    iso_year, iso_week, _ = parsed.isocalendar()
    return {
        "reporting_year": iso_year,
        "reporting_week": iso_week,
        "reporting_week_label": f"{iso_year}-W{iso_week:02d}",
    }

def normalize_phone(phone: str):
    """Normalize phone values to a simple E.164-like format."""
    digits = re.sub(r"\D", "", phone or "")
    if not digits:
        return ""
    if len(digits) == 10:
        digits = f"1{digits}"
    return f"+{digits}"

def parse_hours(text: str):
    """Extract a number from a free-form reply like '32', 'about 40', '~35.5'."""
    match = re.search(r'\d+(\.\d+)?', text)
    return float(match.group()) if match else None

def parse_segment_hours(text: str, week_of: str):
    """Parse split-week replies like 'Apr 27-30=32; May 1-3=8' in segment order."""
    segments = reporting_week_segments(week_of)
    if len(segments) <= 1:
        return {}
    assignment_values = re.findall(r"(?:=|:)\s*(\d+(?:\.\d+)?)", text or "")
    if len(assignment_values) >= len(segments):
        return {
            segment["reporting_week_segment"]: float(assignment_values[index])
            for index, segment in enumerate(segments)
        }
    return {}

def utc_now_iso():
    return datetime.now(timezone.utc).isoformat()

def month_bounds(month_value: str):
    parsed = datetime.strptime(month_value, "%Y-%m")
    start = parsed.strftime("%Y-%m-01")
    end = parsed.replace(day=monthrange(parsed.year, parsed.month)[1]).strftime("%Y-%m-%d")
    label = parsed.strftime("%B %Y")
    return start, end, label

def format_export_timestamp(value: str):
    """Format a stored UTC timestamp for the weekly Excel/CSV report."""
    if not value:
        return ""
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return value
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    central = parsed.astimezone(CENTRAL_TIME)
    return central.strftime("%m/%d/%Y %I:%M %p %Z")

def format_export_phone(value: str):
    """Keep phone numbers readable when the CSV is opened directly in Excel."""
    phone = normalize_phone(value)
    return f'="{phone}"' if phone else ""

def write_response_csv(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Week Of",
        "Reporting Year",
        "Reporting Week",
        "Reporting Week Label",
        "Consultant Name",
        "Phone Number",
        "Billable Hours",
        "Raw Reply",
        "Received At (Central Time)",
    ])
    for row in rows:
        writer.writerow([
            row["week_of"],
            row["reporting_year"],
            row["reporting_week"],
            row["reporting_week_label"],
            row["name"],
            format_export_phone(row["phone"]),
            row["hours"],
            row["raw_reply"],
            format_export_timestamp(row["received_at"]),
        ])
    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))

def write_weekly_report_csv(rows, week_of: str):
    week_start = datetime.strptime(week_of, "%Y-%m-%d")
    week_end = week_start + timedelta(days=6)
    total_hours = sum(row["hours"] or 0 for row in rows)
    response_count = len(rows)
    average_hours = total_hours / response_count if response_count else 0

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Billable Hours Weekly Report"])
    writer.writerow(["Week Start Date", week_of])
    writer.writerow(["Week End Date", week_end.strftime("%Y-%m-%d")])
    writer.writerow(["Reporting Period", format_report_period(week_of)])
    writer.writerow(["Reporting Segments", reporting_segment_summary(week_of)])
    writer.writerow(["Responses Received", response_count])
    writer.writerow(["Total Billable Hours", total_hours])
    writer.writerow(["Average Hours", f"{average_hours:.1f}" if response_count else ""])
    writer.writerow(["Generated At (Central Time)", format_export_timestamp(utc_now_iso())])
    writer.writerow([])
    writer.writerow([
        "Consultant Name",
        "Phone Number",
        "Billable Hours",
        "Raw Reply",
        "Received At (Central Time)",
        "Week Of",
        "Reporting Week Label",
        "Reporting Segments",
    ])
    for row in rows:
        writer.writerow([
            row["name"],
            format_export_phone(row["phone"]),
            row["hours"],
            row["raw_reply"],
            format_export_timestamp(row["received_at"]),
            row["week_of"],
            row["reporting_week_label"],
            reporting_segment_summary(row["week_of"]),
        ])
    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))

def write_monthly_report_csv(rows, month_label: str):
    total_allocated_hours = sum(row["allocated_billable_hours"] or 0 for row in rows)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Billable Hours Monthly Report"])
    writer.writerow(["Report Month", month_label])
    writer.writerow(["Responses/Segments", len(rows)])
    writer.writerow(["Allocated Billable Hours", total_allocated_hours])
    writer.writerow(["Generated At (Central Time)", format_export_timestamp(utc_now_iso())])
    writer.writerow([])
    writer.writerow([
        "Report Month",
        "Week Start Date",
        "Week End Date",
        "Reporting Period",
        "Reporting Week Segment",
        "Segment Start Date",
        "Segment End Date",
        "Segment Days",
        "Weekly Billable Hours",
        "Allocated Billable Hours",
        "Allocation Method",
        "Consultant Name",
        "Phone Number",
        "Raw Reply",
        "Received At (Central Time)",
    ])
    for row in rows:
        writer.writerow([
            month_label,
            row["week_of"],
            row["week_end"],
            format_report_period(row["week_of"]),
            row["reporting_week_segment"],
            row["segment_start"],
            row["segment_end"],
            row["segment_days"],
            row["weekly_billable_hours"],
            row["allocated_billable_hours"],
            row["allocation_method"],
            row["name"],
            format_export_phone(row["phone"]),
            row["raw_reply"],
            format_export_timestamp(row["received_at"]),
        ])
    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))

def query_response_export_rows(where_sql="", params=()):
    received_at = utc_now_iso()
    with get_db() as db:
        return db.execute(
            f"""
            SELECT r.week_of,
                   r.reporting_year,
                   r.reporting_week,
                   r.reporting_week_label,
                   c.name,
                   c.phone,
                   r.hours,
                   r.raw_reply,
                   r.received_at
            FROM responses r JOIN consultants c ON c.id = r.consultant_id
            {where_sql}
            ORDER BY r.week_of DESC, c.name
            """,
            params,
        ).fetchall()

def format_report_period(week_of: str):
    start = datetime.strptime(week_of, "%Y-%m-%d")
    end = start + timedelta(days=6)
    if start.year == end.year:
        if start.month == end.month:
            return f"{start.strftime('%b')} {start.day}-{end.day}, {start.year}"
        return f"{start.strftime('%b')} {start.day}-{end.strftime('%b')} {end.day}, {start.year}"
    return f"{start.strftime('%b')} {start.day}, {start.year}-{end.strftime('%b')} {end.day}, {end.year}"

def format_segment_period(start: datetime, end: datetime):
    if start.year == end.year:
        if start.month == end.month:
            return f"{start.strftime('%b')} {start.day}-{end.day}, {start.year}"
        return f"{start.strftime('%b')} {start.day}-{end.strftime('%b')} {end.day}, {start.year}"
    return f"{start.strftime('%b')} {start.day}, {start.year}-{end.strftime('%b')} {end.day}, {end.year}"

def reporting_week_segments(week_of: str):
    """Split a reporting week into month-specific segments when it crosses a month."""
    week_start = datetime.strptime(week_of, "%Y-%m-%d")
    week_end = week_start + timedelta(days=6)
    base_label = reporting_period_parts(week_of)["reporting_week_label"]
    segments = []
    cursor = week_start

    while cursor <= week_end:
        month_last_day = monthrange(cursor.year, cursor.month)[1]
        segment_end = min(week_end, cursor.replace(day=month_last_day))
        segments.append({
            "segment_start": cursor.strftime("%Y-%m-%d"),
            "segment_end": segment_end.strftime("%Y-%m-%d"),
            "segment_month": cursor.strftime("%Y-%m"),
            "segment_days": (segment_end - cursor).days + 1,
            "segment_period": format_segment_period(cursor, segment_end),
        })
        cursor = segment_end + timedelta(days=1)

    for index, segment in enumerate(segments):
        suffix = chr(ord("A") + index) if len(segments) > 1 else ""
        segment["reporting_week_segment"] = f"{base_label}{suffix}"

    return segments

def reporting_segment_summary(week_of: str):
    return "; ".join(
        f"{segment['reporting_week_segment']}: {segment['segment_period']}"
        for segment in reporting_week_segments(week_of)
    )

def format_segment_reply_label(segment):
    start = datetime.strptime(segment["segment_start"], "%Y-%m-%d")
    end = datetime.strptime(segment["segment_end"], "%Y-%m-%d")
    if start.month == end.month:
        return f"{start.strftime('%b')} {start.day}-{end.day}"
    return f"{start.strftime('%b')} {start.day}-{end.strftime('%b')} {end.day}"

def split_week_reply_template(week_of: str):
    segments = reporting_week_segments(week_of)
    if len(segments) <= 1:
        return ""
    return "; ".join(f"{format_segment_reply_label(segment)}=__" for segment in segments)

def weekly_prompt_message(name: str, week_of: str):
    period = format_report_period(week_of)
    template = split_week_reply_template(week_of)
    if template:
        return (
            f"Hi {name}! Please send billable hours for {period}. "
            f"This week crosses months, so reply with: {template}. "
            "Example: Apr 27-30=32; May 1-3=8. Reply STOP to opt out or HELP for help."
        )
    return (
        f"Hi {name}! How many billable hours will you log for {period}? "
        "Reply with just a number, e.g. 40. Reply STOP to opt out or HELP for help."
    )

def get_monthly_report_rows(month_value: str):
    start, end, label = month_bounds(month_value)
    with get_db() as db:
        rows = db.execute(
            """
            SELECT r.week_of,
                   r.consultant_id,
                   date(r.week_of, '+6 days') AS week_end,
                   r.reporting_year,
                   r.reporting_week,
                   r.reporting_week_label,
                   c.name,
                   c.phone,
                   r.hours,
                   r.raw_reply,
                   r.received_at
            FROM responses r
            JOIN consultants c ON c.id = r.consultant_id
            WHERE r.week_of <= ?
              AND date(r.week_of, '+6 days') >= ?
            ORDER BY r.week_of, c.name
            """,
            (end, start),
        ).fetchall()
        segment_rows = db.execute(
            """
            SELECT consultant_id, week_of, segment_label, hours
            FROM response_segments
            WHERE segment_month = ?
            """,
            (month_value,),
        ).fetchall()
    segment_hours_by_key = {
        (row["consultant_id"], row["week_of"], row["segment_label"]): row["hours"]
        for row in segment_rows
    }
    expanded_rows = []
    for row in rows:
        row_dict = dict(row)
        hours = row_dict["hours"] or 0
        for segment in reporting_week_segments(row_dict["week_of"]):
            if segment["segment_month"] != month_value:
                continue
            exact_segment_hours = segment_hours_by_key.get((
                row_dict["consultant_id"],
                row_dict["week_of"],
                segment["reporting_week_segment"],
            ))
            allocated_hours = (
                exact_segment_hours
                if exact_segment_hours is not None
                else round(hours * segment["segment_days"] / 7, 2)
            )
            expanded_row = {
                **row_dict,
                **segment,
                "weekly_billable_hours": hours,
                "allocated_billable_hours": allocated_hours,
                "allocation_method": "Consultant split reply" if exact_segment_hours is not None else "Prorated weekly total",
            }
            expanded_rows.append(expanded_row)
    return expanded_rows, label

def autosize_worksheet(sheet):
    for column_cells in sheet.columns:
        width = 12
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 42))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

def build_monthly_workbook(month_value: str):
    rows, month_label = get_monthly_report_rows(month_value)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    detail = workbook.create_sheet("Detail")
    charts = workbook.create_sheet("Charts")

    title_fill = PatternFill("solid", fgColor="001844")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="001844")

    total_hours = sum(row["allocated_billable_hours"] or 0 for row in rows)
    consultants = {row["name"] for row in rows}
    week_segments = {row["reporting_week_segment"] for row in rows}

    summary["A1"] = f"{month_label} Billability Report"
    summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = title_fill
    summary.merge_cells("A1:D1")
    summary.append([])
    summary.append(["Metric", "Value"])
    summary.append(["Allocated Billable Hours", total_hours])
    summary.append(["Responses/Segments", len(rows)])
    summary.append(["Consultants Reporting", len(consultants)])
    summary.append(["Reporting Week Segments Included", len(week_segments)])

    for cell in summary[3]:
        cell.fill = header_fill
        cell.font = header_font

    by_consultant = {}
    by_week = {}
    week_segment_meta = {}
    for row in rows:
        by_consultant[row["name"]] = by_consultant.get(row["name"], 0) + (row["allocated_billable_hours"] or 0)
        by_week[row["reporting_week_segment"]] = by_week.get(row["reporting_week_segment"], 0) + (row["allocated_billable_hours"] or 0)
        week_segment_meta[row["reporting_week_segment"]] = row["segment_period"]

    summary.append([])
    summary.append(["Week Segment", "Segment Period", "Allocated Billable Hours"])
    for cell in summary[9]:
        cell.fill = header_fill
        cell.font = header_font
    for segment_label, hours in sorted(by_week.items()):
        summary.append([segment_label, week_segment_meta.get(segment_label, ""), hours])

    consultant_header_row = summary.max_row + 2
    summary.cell(row=consultant_header_row, column=1, value="Consultant Name")
    summary.cell(row=consultant_header_row, column=2, value="Allocated Billable Hours")
    for cell in summary[consultant_header_row][0:2]:
        cell.fill = header_fill
        cell.font = header_font
    for name, hours in sorted(by_consultant.items()):
        summary.append([name, hours])

    detail_headers = [
        "Report Month",
        "Week Start Date",
        "Week End Date",
        "Reporting Period",
        "Reporting Week Segment",
        "Segment Start Date",
        "Segment End Date",
        "Segment Days",
        "Reporting Year",
        "Reporting Week",
        "Reporting Week Label",
        "Consultant Name",
        "Phone Number",
        "Weekly Billable Hours",
        "Allocated Billable Hours",
        "Allocation Method",
        "Raw Reply",
        "Received At (Central Time)",
    ]
    detail.append(detail_headers)
    for cell in detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row in rows:
        detail.append([
            month_label,
            row["week_of"],
            row["week_end"],
            format_report_period(row["week_of"]),
            row["reporting_week_segment"],
            row["segment_start"],
            row["segment_end"],
            row["segment_days"],
            row["reporting_year"],
            row["reporting_week"],
            row["reporting_week_label"],
            row["name"],
            normalize_phone(row["phone"]),
            row["weekly_billable_hours"],
            row["allocated_billable_hours"],
            row["allocation_method"],
            row["raw_reply"],
            format_export_timestamp(row["received_at"]),
        ])

    detail.freeze_panes = "A2"
    summary.freeze_panes = "A4"

    charts["A1"] = f"{month_label} Billability Report - Visual Summary"
    charts["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    charts["A1"].fill = title_fill
    charts.merge_cells("A1:H1")

    charts["A3"] = "Consultant Name"
    charts["B3"] = "Allocated Billable Hours"
    for cell in charts[3][0:2]:
        cell.fill = header_fill
        cell.font = header_font
    chart_row = 4
    for name, hours in sorted(by_consultant.items()):
        charts.cell(row=chart_row, column=1, value=name)
        charts.cell(row=chart_row, column=2, value=hours)
        chart_row += 1

    charts["D3"] = "Reporting Week Segment"
    charts["E3"] = "Allocated Billable Hours"
    for cell in charts[3][3:5]:
        cell.fill = header_fill
        cell.font = header_font
    week_row = 4
    for week_label, hours in sorted(by_week.items()):
        charts.cell(row=week_row, column=4, value=week_label)
        charts.cell(row=week_row, column=5, value=hours)
        week_row += 1

    if by_consultant:
        consultant_chart = BarChart()
        consultant_chart.type = "bar"
        consultant_chart.title = "Allocated Hours by Consultant"
        consultant_chart.y_axis.title = "Consultant"
        consultant_chart.x_axis.title = "Billable Hours"
        data = Reference(charts, min_col=2, min_row=3, max_row=chart_row - 1)
        categories = Reference(charts, min_col=1, min_row=4, max_row=chart_row - 1)
        consultant_chart.add_data(data, titles_from_data=True)
        consultant_chart.set_categories(categories)
        consultant_chart.height = 8
        consultant_chart.width = 14
        charts.add_chart(consultant_chart, "A10")

    if by_week:
        week_chart = BarChart()
        week_chart.type = "col"
        week_chart.title = "Allocated Hours by Reporting Week Segment"
        week_chart.y_axis.title = "Billable Hours"
        week_chart.x_axis.title = "Reporting Week Segment"
        data = Reference(charts, min_col=5, min_row=3, max_row=week_row - 1)
        categories = Reference(charts, min_col=4, min_row=4, max_row=week_row - 1)
        week_chart.add_data(data, titles_from_data=True)
        week_chart.set_categories(categories)
        week_chart.height = 8
        week_chart.width = 14
        charts.add_chart(week_chart, "J10")

    autosize_worksheet(summary)
    autosize_worksheet(detail)
    autosize_worksheet(charts)
    return workbook, month_label

def monthly_report_bytes(month_value: str):
    workbook, _ = build_monthly_workbook(month_value)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def save_monthly_excel_report(month_value: str = None):
    month_value = month_value or datetime.now(timezone.utc).strftime("%Y-%m")
    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(REPORT_OUTPUT_DIR, f"billable_hours_{month_value}.xlsx")
    workbook, _ = build_monthly_workbook(month_value)
    workbook.save(output_path)
    return output_path

def wants_json_response():
    return request.path.startswith("/api/")

def require_basic_auth():
    return Response(
        "Authentication required",
        401,
        {"WWW-Authenticate": 'Basic realm="Billable Hours Tracker"'},
    )

def is_authorized_request():
    if not ADMIN_PASSWORD:
        return True
    auth = request.authorization
    if not auth:
        return False
    return (
        secrets.compare_digest(auth.username or "", ADMIN_USERNAME)
        and secrets.compare_digest(auth.password or "", ADMIN_PASSWORD)
    )

@app.before_request
def protect_dashboard_and_api():
    public_endpoints = {
        "static",
        "sms_reply",
        "healthz",
        "sms_terms",
        "sms_privacy",
        "sms_opt_in",
        "sms_signup",
    }
    if request.endpoint in public_endpoints:
        return None
    if is_authorized_request():
        return None
    if wants_json_response():
        return jsonify({"error": "authentication required"}), 401, {
            "WWW-Authenticate": 'Basic realm="Billable Hours Tracker"'
        }
    return require_basic_auth()

def is_valid_twilio_request():
    if not VALIDATE_TWILIO_WEBHOOKS or app.config.get("TESTING"):
        return True
    if not TWILIO_AUTH_TOKEN:
        return False
    signature = request.headers.get("X-Twilio-Signature", "")
    validator = RequestValidator(TWILIO_AUTH_TOKEN)
    if PUBLIC_BASE_URL:
        path = request.path
        if request.query_string:
            path = f"{path}?{request.query_string.decode()}"
        url = f"{PUBLIC_BASE_URL.rstrip('/')}{path}"
    else:
        url = request.url
    return validator.validate(url, request.form, signature)

# ─── Database ──────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def ensure_week_metadata_columns(db, table_name: str):
    columns = {row["name"] for row in db.execute(f"PRAGMA table_info({table_name})").fetchall()}
    desired_columns = {
        "reporting_year": "INTEGER",
        "reporting_week": "INTEGER",
        "reporting_week_label": "TEXT",
    }
    for column_name, column_type in desired_columns.items():
        if column_name not in columns:
            db.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")

def ensure_message_log_error_code_column(db):
    columns = {row["name"] for row in db.execute("PRAGMA table_info(message_log)").fetchall()}
    if "error_code" not in columns:
        db.execute("ALTER TABLE message_log ADD COLUMN error_code TEXT")

def backfill_week_metadata(db, table_name: str):
    rows = db.execute(
        f"""
        SELECT id, week_of
        FROM {table_name}
        WHERE reporting_year IS NULL
           OR reporting_week IS NULL
           OR reporting_week_label IS NULL
        """
    ).fetchall()
    for row in rows:
        parts = reporting_period_parts(row["week_of"])
        db.execute(
            f"""
            UPDATE {table_name}
            SET reporting_year = ?,
                reporting_week = ?,
                reporting_week_label = ?
            WHERE id = ?
            """,
            (
                parts["reporting_year"],
                parts["reporting_week"],
                parts["reporting_week_label"],
                row["id"],
            )
        )

def init_db():
    with get_db() as db:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS consultants (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                name      TEXT NOT NULL,
                phone     TEXT NOT NULL UNIQUE
            );

            CREATE TABLE IF NOT EXISTS responses (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                consultant_id INTEGER NOT NULL,
                week_of       TEXT NOT NULL,   -- ISO date of Monday
                reporting_year INTEGER,
                reporting_week INTEGER,
                reporting_week_label TEXT,
                hours         REAL,
                raw_reply     TEXT,
                received_at   TEXT NOT NULL,
                FOREIGN KEY (consultant_id) REFERENCES consultants(id),
                UNIQUE(consultant_id, week_of)   -- one entry per person per week
            );

            CREATE TABLE IF NOT EXISTS response_segments (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                consultant_id   INTEGER NOT NULL,
                week_of         TEXT NOT NULL,
                segment_label   TEXT NOT NULL,
                segment_start   TEXT NOT NULL,
                segment_end     TEXT NOT NULL,
                segment_month   TEXT NOT NULL,
                hours           REAL NOT NULL,
                raw_reply       TEXT,
                received_at     TEXT NOT NULL,
                FOREIGN KEY (consultant_id) REFERENCES consultants(id),
                UNIQUE(consultant_id, week_of, segment_label)
            );

            CREATE TABLE IF NOT EXISTS sent_log (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                consultant_id INTEGER NOT NULL,
                week_of       TEXT NOT NULL,
                reporting_year INTEGER,
                reporting_week INTEGER,
                reporting_week_label TEXT,
                sent_at       TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS message_log (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                consultant_id INTEGER NOT NULL,
                week_of       TEXT NOT NULL,
                reporting_year INTEGER,
                reporting_week INTEGER,
                reporting_week_label TEXT,
                message_type  TEXT NOT NULL,
                body          TEXT NOT NULL,
                status        TEXT NOT NULL,
                error_message TEXT,
                error_code    TEXT,
                twilio_sid    TEXT,
                sent_at       TEXT NOT NULL,
                FOREIGN KEY (consultant_id) REFERENCES consultants(id)
            );

            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS opt_in_log (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                name          TEXT NOT NULL,
                phone         TEXT NOT NULL,
                consent_text  TEXT NOT NULL,
                source        TEXT NOT NULL,
                ip_address    TEXT,
                user_agent    TEXT,
                created_at    TEXT NOT NULL
            );
        """)
        ensure_week_metadata_columns(db, "responses")
        ensure_week_metadata_columns(db, "sent_log")
        ensure_week_metadata_columns(db, "message_log")
        ensure_message_log_error_code_column(db)
        consultants = db.execute("SELECT id, phone FROM consultants").fetchall()
        for consultant in consultants:
            normalized_phone = normalize_phone(consultant["phone"])
            if normalized_phone and normalized_phone != consultant["phone"]:
                db.execute(
                    "UPDATE consultants SET phone = ? WHERE id = ?",
                    (normalized_phone, consultant["id"])
                )
        backfill_week_metadata(db, "responses")
        backfill_week_metadata(db, "sent_log")
        backfill_week_metadata(db, "message_log")

init_db()

def get_reporting_cycle_date():
    with get_db() as db:
        row = db.execute(
            "SELECT value FROM app_settings WHERE key = ?",
            ("reporting_cycle_date",)
        ).fetchone()
    return row["value"] if row else None

def set_reporting_cycle_date(value: str):
    normalized_value = validate_reporting_cycle_date(value)
    with get_db() as db:
        db.execute(
            """
            INSERT INTO app_settings (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            ("reporting_cycle_date", normalized_value)
        )
    return normalized_value

def replace_response_segments(consultant_id: int, week_of: str, segment_hours: dict, raw_reply: str, received_at: str):
    segments = {
        segment["reporting_week_segment"]: segment
        for segment in reporting_week_segments(week_of)
    }
    with get_db() as db:
        db.execute(
            "DELETE FROM response_segments WHERE consultant_id = ? AND week_of = ?",
            (consultant_id, week_of),
        )
        for segment_label, hours in segment_hours.items():
            segment = segments.get(segment_label)
            if not segment:
                continue
            db.execute(
                """
                INSERT INTO response_segments
                    (consultant_id, week_of, segment_label, segment_start, segment_end, segment_month, hours, raw_reply, received_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    consultant_id,
                    week_of,
                    segment_label,
                    segment["segment_start"],
                    segment["segment_end"],
                    segment["segment_month"],
                    hours,
                    raw_reply,
                    received_at,
                ),
            )

def send_sms(to: str, body: str):
    if not TWILIO_ACCOUNT_SID or not TWILIO_AUTH_TOKEN:
        raise RuntimeError("Twilio is not configured. Set TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN.")
    if not TWILIO_PHONE and not TWILIO_MESSAGING_SERVICE_SID:
        raise RuntimeError("Twilio is not configured. Set TWILIO_PHONE or TWILIO_MESSAGING_SERVICE_SID.")

    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
    message_args = {"to": to, "body": body}
    if TWILIO_MESSAGING_SERVICE_SID:
        message_args["messaging_service_sid"] = TWILIO_MESSAGING_SERVICE_SID
    else:
        message_args["from_"] = TWILIO_PHONE
    return client.messages.create(**message_args)

def default_reminder_message(name: str, week_of: str):
    template = split_week_reply_template(week_of)
    if template:
        return (
            f"Hi {name}, reminder to send billable hours for {format_report_period(week_of)}. "
            f"Reply with: {template}. Reply STOP to opt out or HELP for help."
        )
    return (
        f"Hi {name}, quick follow-up on your billable hours forecast for {format_report_period(week_of)}. "
        "Please reply with the number of hours you expect to log. Reply STOP to opt out or HELP for help."
    )

def log_message_attempt(consultant_id: int, week_of: str, message_type: str, body: str, status: str,
                        twilio_sid: str = None, error_message: str = None, error_code: str = None,
                        sent_at: str = None):
    sent_at = sent_at or utc_now_iso()
    period = reporting_period_parts(week_of)
    with get_db() as db:
        db.execute(
            """
            INSERT INTO message_log
                (
                    consultant_id, week_of, reporting_year, reporting_week, reporting_week_label,
                    message_type, body, status, error_message, error_code, twilio_sid, sent_at
                )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                consultant_id,
                week_of,
                period["reporting_year"],
                period["reporting_week"],
                period["reporting_week_label"],
                message_type,
                body,
                status,
                error_message,
                error_code,
                twilio_sid,
                sent_at,
            )
        )

def send_consultant_message(consultant, body: str, week_of: str, message_type: str):
    try:
        msg = send_sms(consultant["phone"], body)
        sent_at = utc_now_iso()
        log_message_attempt(
            consultant["id"],
            week_of,
            message_type,
            body,
            "sent",
            twilio_sid=msg.sid,
            sent_at=sent_at
        )
        if message_type == "weekly_prompt":
            period = reporting_period_parts(week_of)
            with get_db() as db:
                db.execute(
                    """
                    INSERT OR IGNORE INTO sent_log
                        (consultant_id, week_of, reporting_year, reporting_week, reporting_week_label, sent_at)
                    VALUES (?,?,?,?,?,?)
                    """,
                    (
                        consultant["id"],
                        week_of,
                        period["reporting_year"],
                        period["reporting_week"],
                        period["reporting_week_label"],
                        sent_at,
                    )
                )
        return {"ok": True, "sid": msg.sid}
    except TwilioRestException as exc:
        error_message = exc.msg
        error_code = str(exc.code) if exc.code else None
    except Exception as exc:
        error_message = str(exc)
        error_code = None

    log_message_attempt(
        consultant["id"],
        week_of,
        message_type,
        body,
        "failed",
        error_message=error_message,
        error_code=error_code
    )
    return {"ok": False, "error": error_message}

def get_consultant_by_id(cid: int):
    with get_db() as db:
        consultant = db.execute("SELECT * FROM consultants WHERE id = ?", (cid,)).fetchone()
    return consultant

def get_weekly_reporting_status(week_of: str):
    with get_db() as db:
        consultants = db.execute("SELECT * FROM consultants ORDER BY name").fetchall()
        responses = db.execute(
            """
            SELECT consultant_id, hours, raw_reply, received_at
            FROM responses
            WHERE week_of = ?
            """,
            (week_of,)
        ).fetchall()
        messages = db.execute(
            """
            SELECT ml.*
            FROM message_log ml
            JOIN (
                SELECT consultant_id, MAX(id) AS max_id
                FROM message_log
                WHERE week_of = ?
                GROUP BY consultant_id
            ) latest ON latest.max_id = ml.id
            """,
            (week_of,)
        ).fetchall()

    responses_by_consultant = {row["consultant_id"]: dict(row) for row in responses}
    messages_by_consultant = {row["consultant_id"]: dict(row) for row in messages}
    items = []
    for consultant in consultants:
        response = responses_by_consultant.get(consultant["id"])
        message = messages_by_consultant.get(consultant["id"])
        if response:
            status_key = "replied"
            status_label = "Replied"
        elif message and message["status"] == "failed":
            status_key = "failed"
            status_label = "Failed to send"
        elif message and message["message_type"] in ("reminder", "manual_follow_up"):
            status_key = "needs_follow_up"
            status_label = "Needs follow-up"
        elif message:
            status_key = "pending"
            status_label = "Pending"
        else:
            status_key = "pending"
            status_label = "Pending"

        items.append({
            "consultant_id": consultant["id"],
            "name": consultant["name"],
            "phone": consultant["phone"],
            "week_of": week_of,
            "hours": response["hours"] if response else None,
            "raw_reply": response["raw_reply"] if response else None,
            "received_at": response["received_at"] if response else None,
            "status": status_key,
            "status_label": status_label,
            "last_message_type": message["message_type"] if message else None,
            "last_message_status": message["status"] if message else None,
            "last_message_at": message["sent_at"] if message else None,
            "last_error": message["error_message"] if message else None,
            "last_error_code": message["error_code"] if message else None,
        })
    return items

def get_history_for_consultant(consultant_id: int, week_of: str):
    with get_db() as db:
        message_rows = db.execute(
            """
            SELECT id, message_type, body, status, error_message, error_code, sent_at, twilio_sid
            FROM message_log
            WHERE consultant_id = ? AND week_of = ?
            ORDER BY sent_at DESC, id DESC
            """,
            (consultant_id, week_of)
        ).fetchall()
        response_rows = db.execute(
            """
            SELECT id, hours, raw_reply, received_at
            FROM responses
            WHERE consultant_id = ? AND week_of = ?
            ORDER BY received_at DESC, id DESC
            """,
            (consultant_id, week_of)
        ).fetchall()

    items = [
        {
            "kind": "message",
            "label": MESSAGE_TYPE_LABELS.get(row["message_type"], row["message_type"]),
            "message_type": row["message_type"],
            "body": row["body"],
            "status": row["status"],
            "error_message": row["error_message"],
            "error_code": row["error_code"],
            "timestamp": row["sent_at"],
            "twilio_sid": row["twilio_sid"],
        }
        for row in message_rows
    ]
    items.extend(
        {
            "kind": "response",
            "label": MESSAGE_TYPE_LABELS["response"],
            "hours": row["hours"],
            "body": row["raw_reply"],
            "status": "received",
            "timestamp": row["received_at"],
        }
        for row in response_rows
    )
    return sorted(items, key=lambda item: item["timestamp"], reverse=True)

# ─── Weekly job ────────────────────────────────────────────────────────────────

def send_weekly_texts():
    """Runs every Monday morning. Sends the question to all consultants."""
    week_of = current_week_monday()
    results = {"sent": 0, "failed": []}
    with get_db() as db:
        consultants = db.execute("SELECT * FROM consultants").fetchall()
    for c in consultants:
        body = weekly_prompt_message(c["name"], week_of)
        result = send_consultant_message(c, body, week_of, "weekly_prompt")
        if result["ok"]:
            results["sent"] += 1
        else:
            message = f"{c['name']} ({c['phone']}): {result['error']}"
            print(f"Failed to text {message}")
            results["failed"].append(message)
    return results

def run_monthly_excel_job():
    """Generate this month's Excel report on the final calendar day."""
    month_value = datetime.now(timezone.utc).strftime("%Y-%m")
    try:
        output_path = save_monthly_excel_report(month_value)
        print(f"Monthly Excel report saved to {output_path}")
        return output_path
    except Exception as exc:
        if sentry_sdk:
            sentry_sdk.capture_exception(exc)
        raise

# Schedules are expressed in UTC for deployment consistency.
scheduler = BackgroundScheduler(timezone=timezone.utc)
scheduler.add_job(send_weekly_texts, "cron", day_of_week="mon", hour=WEEKLY_PROMPT_UTC_HOUR, minute=0)
scheduler.add_job(run_monthly_excel_job, "cron", day="last", hour=MONTHLY_REPORT_UTC_HOUR, minute=0)
scheduler.start()

# ─── Twilio webhook (handles inbound replies) ──────────────────────────────────

@app.route("/sms", methods=["POST"])
def sms_reply():
    if not is_valid_twilio_request():
        return Response("Invalid Twilio signature", status=403)

    from_number = normalize_phone(request.form.get("From", "").strip())
    body        = request.form.get("Body", "").strip()
    week_of     = current_week_monday()
    segment_hours = parse_segment_hours(body, week_of)
    hours       = sum(segment_hours.values()) if segment_hours else parse_hours(body)

    resp = MessagingResponse()

    with get_db() as db:
        consultant = db.execute(
            "SELECT * FROM consultants WHERE phone = ?", (from_number,)
        ).fetchone()

    if not consultant:
        resp.message("Sorry, your number isn't registered. Ask your manager to add you.")
        return str(resp)

    received_at = utc_now_iso()
    with get_db() as db:
        period = reporting_period_parts(week_of)
        db.execute(
            """INSERT INTO responses
               (consultant_id, week_of, reporting_year, reporting_week, reporting_week_label, hours, raw_reply, received_at)
               VALUES (?,?,?,?,?,?,?,?)
               ON CONFLICT(consultant_id, week_of) DO UPDATE SET
                   reporting_year=excluded.reporting_year,
                   reporting_week=excluded.reporting_week,
                   reporting_week_label=excluded.reporting_week_label,
                   hours=excluded.hours,
                   raw_reply=excluded.raw_reply,
                   received_at=excluded.received_at""",
            (
                consultant["id"],
                week_of,
                period["reporting_year"],
                period["reporting_week"],
                period["reporting_week_label"],
                hours,
                body,
                received_at,
            )
        )
    if segment_hours:
        replace_response_segments(consultant["id"], week_of, segment_hours, body, received_at)
    log_message_attempt(
        consultant["id"],
        week_of,
        "response",
        body,
        "received",
        sent_at=received_at
    )

    if hours is not None and segment_hours:
        detail = ", ".join(f"{label}: {value:g}" for label, value in segment_hours.items())
        resp.message(f"Got it - {hours:g} total hours logged for {week_of} ({detail}). Thanks!")
    elif hours is not None:
        resp.message(f"Got it - {hours:.0f} hours logged for the week of {week_of}. Thanks!")
    else:
        template = split_week_reply_template(week_of)
        if template:
            resp.message(f"Hmm, I did not catch the split hours. Please reply like: {template}.")
        else:
            resp.message("Hmm, I did not catch a number. Please reply with just digits, e.g. 40.")

    return str(resp)

# ─── REST API ─────────────────────────────────────────────────────────────────

@app.route("/api/consultants", methods=["GET"])
def list_consultants():
    with get_db() as db:
        rows = db.execute("SELECT * FROM consultants ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/consultants", methods=["POST"])
def add_consultant():
    data = request.json
    name  = data.get("name", "").strip()
    phone = normalize_phone(data.get("phone", "").strip())
    if not name or not phone:
        return jsonify({"error": "name and phone required"}), 400
    try:
        with get_db() as db:
            db.execute("INSERT INTO consultants (name, phone) VALUES (?,?)", (name, phone))
    except sqlite3.IntegrityError:
        return jsonify({"error": "consultant already exists for that phone number"}), 409
    return jsonify({"ok": True}), 201

@app.route("/api/consultants/<int:cid>", methods=["DELETE"])
def remove_consultant(cid):
    with get_db() as db:
        db.execute("DELETE FROM consultants WHERE id=?", (cid,))
    return jsonify({"ok": True})

@app.route("/api/responses", methods=["GET"])
def list_responses():
    weeks = request.args.get("weeks", 8, type=int)
    cutoff = (datetime.now(timezone.utc) - timedelta(weeks=weeks)).strftime("%Y-%m-%d")
    with get_db() as db:
        rows = db.execute("""
            SELECT r.week_of, c.name, r.hours, r.raw_reply, r.received_at
                 , r.reporting_year, r.reporting_week, r.reporting_week_label
            FROM responses r JOIN consultants c ON c.id = r.consultant_id
            WHERE r.week_of >= ?
            ORDER BY r.week_of DESC, c.name
        """, (cutoff,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/reporting-status", methods=["GET"])
def reporting_status():
    week_of = request.args.get("week_of", type=str) or current_week_monday()
    try:
        normalized_week = validate_reporting_cycle_date(week_of)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify(get_weekly_reporting_status(normalized_week))

@app.route("/api/summary", methods=["GET"])
def weekly_summary():
    """Total hours per week + per-consultant breakdown."""
    week_of = request.args.get("week_of", type=str)
    if week_of:
        try:
            normalized_week = validate_reporting_cycle_date(week_of)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        with get_db() as db:
            row = db.execute("""
                SELECT r.week_of,
                       r.reporting_year,
                       r.reporting_week,
                       r.reporting_week_label,
                       COUNT(r.id)    AS responses,
                       SUM(r.hours)   AS total_hours,
                       AVG(r.hours)   AS avg_hours
                FROM responses r
                WHERE r.week_of = ?
                GROUP BY r.week_of
            """, (normalized_week,)).fetchone()
        return jsonify([dict(row)] if row else [])

    weeks = request.args.get("weeks", 8, type=int)
    cutoff = (datetime.now(timezone.utc) - timedelta(weeks=weeks)).strftime("%Y-%m-%d")
    with get_db() as db:
        rows = db.execute("""
            SELECT r.week_of,
                   r.reporting_year,
                   r.reporting_week,
                   r.reporting_week_label,
                   COUNT(r.id)    AS responses,
                   SUM(r.hours)   AS total_hours,
                   AVG(r.hours)   AS avg_hours
            FROM responses r
            WHERE r.week_of >= ?
            GROUP BY r.week_of
            ORDER BY r.week_of DESC
        """, (cutoff,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/reporting-cycle", methods=["GET"])
def get_reporting_cycle():
    cycle_date = current_week_monday()
    return jsonify({"reporting_cycle_date": cycle_date})

@app.route("/api/reporting-cycle", methods=["POST"])
def update_reporting_cycle():
    data = request.json or {}
    try:
        cycle_date = set_reporting_cycle_date(data.get("reporting_cycle_date", "").strip())
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({
        "ok": True,
        "reporting_cycle_date": cycle_date,
        "message": f"Reporting cycle updated to {cycle_date}."
    })

@app.route("/api/send-now", methods=["POST"])
def send_now():
    """Manually trigger the weekly text blast (for testing)."""
    results = send_weekly_texts()
    if results["failed"]:
        details = " | ".join(results["failed"][:3])
        return jsonify({
            "ok": False,
            "message": f"Sent {results['sent']} message(s). Failed: {details}"
        }), 502
    return jsonify({"ok": True, "message": f"Sent {results['sent']} text(s)."})

@app.route("/api/messages/send", methods=["POST"])
def send_follow_up():
    data = request.json or {}
    consultant_id = data.get("consultant_id")
    body = (data.get("body") or "").strip()
    message_type = (data.get("message_type") or "manual_follow_up").strip()
    week_of = data.get("week_of", "").strip() or current_week_monday()

    if not consultant_id:
        return jsonify({"error": "consultant_id is required"}), 400
    if not body:
        return jsonify({"error": "body is required"}), 400

    try:
        week_of = validate_reporting_cycle_date(week_of)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    consultant = get_consultant_by_id(int(consultant_id))
    if not consultant:
        return jsonify({"error": "consultant not found"}), 404

    result = send_consultant_message(consultant, body, week_of, message_type)
    if not result["ok"]:
        return jsonify({"ok": False, "message": result["error"]}), 502
    return jsonify({"ok": True, "message": f"Message sent to {consultant['name']}."})

@app.route("/api/messages/bulk-remind", methods=["POST"])
def bulk_remind():
    data = request.json or {}
    week_of = data.get("week_of", "").strip() or current_week_monday()
    try:
        week_of = validate_reporting_cycle_date(week_of)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    statuses = get_weekly_reporting_status(week_of)
    pending = [item for item in statuses if item["status"] in ("pending", "needs_follow_up", "failed")]
    sent = 0
    failed = []

    for item in pending:
        consultant = get_consultant_by_id(item["consultant_id"])
        body = default_reminder_message(consultant["name"], week_of)
        result = send_consultant_message(consultant, body, week_of, "reminder")
        if result["ok"]:
            sent += 1
        else:
            failed.append(f"{consultant['name']}: {result['error']}")

    if failed:
        details = " | ".join(failed[:3])
        return jsonify({
            "ok": False,
            "message": f"Sent {sent} reminder(s). Failed: {details}"
        }), 502
    return jsonify({"ok": True, "message": f"Sent {sent} reminder(s)."})

@app.route("/api/consultants/<int:cid>/history", methods=["GET"])
def consultant_history(cid):
    week_of = request.args.get("week_of", type=str) or current_week_monday()
    try:
        week_of = validate_reporting_cycle_date(week_of)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    consultant = get_consultant_by_id(cid)
    if not consultant:
        return jsonify({"error": "consultant not found"}), 404

    return jsonify({
        "consultant": {
            "id": consultant["id"],
            "name": consultant["name"],
            "phone": consultant["phone"],
        },
        "week_of": week_of,
        "history": get_history_for_consultant(cid, week_of),
    })

@app.route("/api/export", methods=["GET"])
def export_csv():
    rows = query_response_export_rows()
    return send_file(
        write_response_csv(rows),
        mimetype="text/csv",
        as_attachment=True,
        download_name="billable_hours.csv"
    )

@app.route("/api/export/weekly", methods=["GET"])
def export_weekly_csv():
    week_of = request.args.get("week_of", current_week_monday(), type=str)
    try:
        normalized_week = validate_reporting_cycle_date(week_of)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    rows = query_response_export_rows("WHERE r.week_of = ?", (normalized_week,))
    return send_file(
        write_weekly_report_csv(rows, normalized_week),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"billable_hours_week_{normalized_week}.csv",
    )

@app.route("/api/export/monthly/csv", methods=["GET"])
def export_monthly_csv():
    month_value = request.args.get("month", datetime.now(timezone.utc).strftime("%Y-%m"), type=str)
    try:
        month_bounds(month_value)
    except ValueError:
        return jsonify({"error": "month must use YYYY-MM format"}), 400

    rows, month_label = get_monthly_report_rows(month_value)
    return send_file(
        write_monthly_report_csv(rows, month_label),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"billable_hours_month_{month_value}.csv",
    )

@app.route("/api/export/monthly", methods=["GET"])
def export_monthly_xlsx():
    month_value = request.args.get("month", datetime.now(timezone.utc).strftime("%Y-%m"), type=str)
    try:
        month_bounds(month_value)
    except ValueError:
        return jsonify({"error": "month must use YYYY-MM format"}), 400

    output = monthly_report_bytes(month_value)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"billable_hours_{month_value}.xlsx",
    )

# ─── Dashboard ────────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    return render_template("dashboard.html")

@app.route("/sms-terms")
def sms_terms():
    return render_template("sms_terms.html")

@app.route("/sms-privacy")
def sms_privacy():
    return render_template("sms_privacy.html")

@app.route("/sms-opt-in")
def sms_opt_in():
    return render_template("sms_opt_in.html")

@app.route("/sms-signup", methods=["GET", "POST"])
def sms_signup():
    consent_text = (
        "I agree to receive recurring SMS messages from EPI-USE America, Inc. "
        "for billable-hours forecasting and follow-up reminders. Message frequency varies, "
        "typically weekly. Message and data rates may apply. Reply STOP to opt out or HELP for help."
    )

    if request.method == "GET":
        return render_template("sms_signup.html", consent_text=consent_text, submitted=False)

    name = (request.form.get("name") or "").strip()
    phone = normalize_phone(request.form.get("phone") or "")
    consent = request.form.get("sms_consent") == "yes"

    errors = []
    if not name:
        errors.append("Name is required.")
    if not phone:
        errors.append("Mobile number is required.")

    if errors:
        return render_template(
            "sms_signup.html",
            consent_text=consent_text,
            errors=errors,
            name=name,
            phone=request.form.get("phone") or "",
            submitted=False,
        ), 400

    if not consent:
        return render_template(
            "sms_signup.html",
            consent_text=consent_text,
            submitted=True,
            sms_opted_in=False,
            name=name,
            phone=phone,
        )

    with get_db() as db:
        existing_phone = db.execute("SELECT id FROM consultants WHERE phone = ?", (phone,)).fetchone()
        existing_name = db.execute("SELECT id FROM consultants WHERE lower(name) = lower(?)", (name,)).fetchone()
        if existing_phone:
            errors.append("That mobile number is already registered.")
        if existing_name:
            errors.append("That name is already registered.")
        if errors:
            return render_template(
                "sms_signup.html",
                consent_text=consent_text,
                errors=errors,
                name=name,
                phone=request.form.get("phone") or "",
                submitted=False,
            ), 409

        db.execute("INSERT INTO consultants (name, phone) VALUES (?, ?)", (name, phone))
        db.execute(
            """
            INSERT INTO opt_in_log (name, phone, consent_text, source, ip_address, user_agent, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                phone,
                consent_text,
                "sms_signup",
                request.headers.get("X-Forwarded-For", request.remote_addr),
                request.headers.get("User-Agent"),
                utc_now_iso(),
            ),
        )

    return render_template(
        "sms_signup.html",
        consent_text=consent_text,
        submitted=True,
        sms_opted_in=True,
        name=name,
        phone=phone,
    )

@app.route("/healthz")
def healthz():
    twilio_ready = bool(
        TWILIO_ACCOUNT_SID
        and TWILIO_AUTH_TOKEN
        and (TWILIO_PHONE or TWILIO_MESSAGING_SERVICE_SID)
    )
    return jsonify({
        "ok": True,
        "twilio_configured": twilio_ready,
        "twilio_sender": "messaging_service" if TWILIO_MESSAGING_SERVICE_SID else "phone_number",
        "webhook_validation_enabled": VALIDATE_TWILIO_WEBHOOKS,
        "admin_auth_enabled": bool(ADMIN_PASSWORD),
    })

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=PORT)
