import importlib
import os
import sys
from io import BytesIO
from datetime import datetime, timedelta, timezone
from pathlib import Path


from openpyxl import load_workbook
import pytest


ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="module")
def tracker_app(tmp_path_factory):
    db_path = tmp_path_factory.mktemp("data") / "hours_test.db"
    os.environ["DB_PATH"] = str(db_path)
    os.environ.pop("ADMIN_PASSWORD", None)
    os.environ["VALIDATE_TWILIO_WEBHOOKS"] = "false"
    sys.path.insert(0, str(ROOT))
    sys.modules.pop("app", None)
    module = importlib.import_module("app")
    module.app.config.update(TESTING=True)
    yield module
    module.scheduler.shutdown(wait=False)


def test_parse_hours(tracker_app):
    assert tracker_app.parse_hours("40") == 40.0
    assert tracker_app.parse_hours("about 37.5 this week") == 37.5
    assert tracker_app.parse_hours("no hours yet") is None


def test_health_check(tracker_app):
    previous_password = tracker_app.ADMIN_PASSWORD
    tracker_app.ADMIN_PASSWORD = "secret"
    try:
        response = tracker_app.app.test_client().get("/healthz")
    finally:
        tracker_app.ADMIN_PASSWORD = previous_password

    assert response.status_code == 200
    assert response.json == {"ok": True}


def test_current_week_monday_uses_current_utc_week_without_setting(tracker_app):
    with tracker_app.get_db() as db:
        db.execute("DELETE FROM app_settings WHERE key = ?", ("reporting_cycle_date",))

    today = datetime.now(timezone.utc)
    expected = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    assert tracker_app.current_week_monday() == expected


def test_sms_reply_saves_response_for_registered_consultant(tracker_app):
    client = tracker_app.app.test_client()
    phone = "+15551234567"
    with tracker_app.get_db() as db:
        db.execute("INSERT INTO consultants (name, phone) VALUES (?, ?)", ("Test Consultant", phone))

    response = client.post("/sms", data={"From": phone, "Body": "42 hours"})

    assert response.status_code == 200
    assert b"42 hours" in response.data

    with tracker_app.get_db() as db:
        row = db.execute(
            """
            SELECT c.name, r.hours, r.raw_reply
            FROM responses r
            JOIN consultants c ON c.id = r.consultant_id
            WHERE c.phone = ?
            """,
            (phone,),
        ).fetchone()
    assert dict(row) == {
        "name": "Test Consultant",
        "hours": 42.0,
        "raw_reply": "42 hours",
    }


def test_monthly_export_returns_xlsx(tracker_app):
    client = tracker_app.app.test_client()
    response = client.get("/api/export/monthly?month=2026-05")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.data), read_only=True)
    assert workbook.sheetnames == ["Summary", "Detail", "Charts"]
    assert workbook["Summary"]["A1"].value == "May 2026 Billability Report"
    assert workbook["Detail"]["A1"].value == "Report Month"
    assert workbook["Charts"]["A1"].value == "May 2026 Billability Report - Visual Summary"


def test_weekly_and_monthly_csv_exports_use_selected_period(tracker_app):
    client = tracker_app.app.test_client()

    weekly = client.get("/api/export/weekly?week_of=2026-05-04")
    monthly = client.get("/api/export/monthly/csv?month=2026-05")

    assert weekly.status_code == 200
    assert weekly.mimetype == "text/csv"
    assert b"Billable Hours Weekly Report" in weekly.data
    assert b"Consultant Name,Phone Number,Billable Hours" in weekly.data

    assert monthly.status_code == 200
    assert monthly.mimetype == "text/csv"
    assert b"Consultant Name,Phone Number" in monthly.data


def test_cross_month_week_splits_into_reporting_segments(tracker_app):
    segments = tracker_app.reporting_week_segments("2026-04-27")

    assert segments == [
        {
            "segment_start": "2026-04-27",
            "segment_end": "2026-04-30",
            "segment_month": "2026-04",
            "segment_days": 4,
            "segment_period": "Apr 27-30, 2026",
            "reporting_week_segment": "2026-W18A",
        },
        {
            "segment_start": "2026-05-01",
            "segment_end": "2026-05-03",
            "segment_month": "2026-05",
            "segment_days": 3,
            "segment_period": "May 1-3, 2026",
            "reporting_week_segment": "2026-W18B",
        },
    ]


def test_cross_month_prompt_asks_for_one_split_reply(tracker_app):
    message = tracker_app.weekly_prompt_message("Alex", "2026-04-27")

    assert "Apr 27-30=__" in message
    assert "May 1-3=__" in message
    assert "This week crosses months" in message


def test_sms_reply_stores_exact_split_hours(tracker_app):
    client = tracker_app.app.test_client()
    phone = "+15559870000"
    with tracker_app.get_db() as db:
        db.execute("INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value", ("reporting_cycle_date", "2026-04-27"))
        db.execute("INSERT INTO consultants (name, phone) VALUES (?, ?)", ("Split Reply", phone))

    response = client.post("/sms", data={"From": phone, "Body": "Apr 27-30=40; May 1-3=30"})

    assert response.status_code == 200
    assert b"2026-W18A: 40" in response.data
    assert b"2026-W18B: 30" in response.data

    with tracker_app.get_db() as db:
        weekly = db.execute(
            """
            SELECT hours
            FROM responses r
            JOIN consultants c ON c.id = r.consultant_id
            WHERE c.phone = ? AND r.week_of = ?
            """,
            (phone, "2026-04-27"),
        ).fetchone()
        segments = db.execute(
            """
            SELECT segment_label, segment_start, segment_end, hours
            FROM response_segments rs
            JOIN consultants c ON c.id = rs.consultant_id
            WHERE c.phone = ?
            ORDER BY segment_label
            """,
            (phone,),
        ).fetchall()

    assert weekly["hours"] == 70
    assert [dict(row) for row in segments] == [
        {"segment_label": "2026-W18A", "segment_start": "2026-04-27", "segment_end": "2026-04-30", "hours": 40},
        {"segment_label": "2026-W18B", "segment_start": "2026-05-01", "segment_end": "2026-05-03", "hours": 30},
    ]


def test_monthly_export_prorates_cross_month_week(tracker_app):
    phone = "+15557654321"
    week = "2026-04-27"
    period = tracker_app.reporting_period_parts(week)
    with tracker_app.get_db() as db:
        cursor = db.execute("INSERT INTO consultants (name, phone) VALUES (?, ?)", ("Cross Month", phone))
        db.execute(
            """
            INSERT INTO responses
                (consultant_id, week_of, reporting_year, reporting_week, reporting_week_label, hours, raw_reply, received_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                cursor.lastrowid,
                week,
                period["reporting_year"],
                period["reporting_week"],
                period["reporting_week_label"],
                70,
                "70",
                "2026-05-01T15:00:00+00:00",
            ),
        )

    april_rows, _ = tracker_app.get_monthly_report_rows("2026-04")
    may_rows, _ = tracker_app.get_monthly_report_rows("2026-05")
    april_row = next(row for row in april_rows if row["name"] == "Cross Month")
    may_row = next(row for row in may_rows if row["name"] == "Cross Month")

    assert april_row["reporting_week_segment"] == "2026-W18A"
    assert april_row["segment_days"] == 4
    assert april_row["allocated_billable_hours"] == 40

    assert may_row["reporting_week_segment"] == "2026-W18B"
    assert may_row["segment_days"] == 3
    assert may_row["allocated_billable_hours"] == 30
